#!/usr/bin/env python
# coding: utf-8

# In[191]:


pip install selenium


# In[201]:


import extractPrimerSet as ps
import openpyxl


# In[202]:


from selenium import webdriver


# In[203]:


#preset drivers
driver, driver2, driver3, driver4 = ps.presetDrivers()


# In[181]:


ps.preset_Primer3(driver3)


# In[182]:


##############preset files##############
input_name = 'amplicon_primer_ALS.xlsx'
output_name = 'amplicon_primer_ALS_order.xlsx'
date = '2021-07-28'
########################################

#open input file
load_wb = openpyxl.load_workbook(input_name, data_only = True)
load_ws = load_wb['list']

write_wb = openpyxl.Workbook()
dest_filename = output_name
write_ws = write_wb.active
write_ws.title = 'order'
header = [date,None,'Name','Chrom','Pos','insert size','seq (Forward/backward)','Characteristics','Adaptor','Order sequence']
write_ws.append(header)


# In[167]:


geneLoc_list = []
for row in load_ws.rows:
    if (row[0].value != None and row[0].value != 'No'):
        geneLoc_list.append(row[2].value)
        

for i in range(len(geneLoc_list)):
    geneLoc = geneLoc_list[i]
    extractPrimer = ps.extractPrimer(geneLoc)
    #returns (True, (left_strt, right_end), insert_size, (F_primer,R_primer), (p1_char, p2_char))
    if extractPrimer[0]: 
        chrom_loc = geneLoc.split(":")[0]
        write_ws.append([i, geneLoc, geneLoc+":F", chrom_loc,extractPrimer[1][0], extractPrimer[2],extractPrimer[3][0], extractPrimer[4][0]])
        write_ws.append(["", "", geneLoc+":R", chrom_loc, extractPrimer[1][1], "",extractPrimer[3][1], extractPrimer[4][1]])
    else:
        write_ws.append ([i, geneLoc,"no good primer"])
        write_ws.append([])

write_wb.save(output_name)


# In[ ]:


###read file###
output_wb = openpyxl.load_workbook(output_name, data_only = True)
output_ws = output_wb['order']

for row in output_ws:
    print("%s\t%s" % (row[0].value, row[1].value))
    row_num += 1

for i in output_ws:
    print (row_num)


# In[ ]:


#clear worksheet
write_wb.remove(write_ws)
write_wb.save(output_name)


# In[ ]:




